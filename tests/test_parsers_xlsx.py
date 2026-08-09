"""Tests for scraper/parsers/xlsx.py: XLSX bond enrichment parsing."""

from __future__ import annotations

import os
from datetime import date, datetime
from decimal import Decimal

import pytest

from scraper.parsers.xlsx import (
    XLSX_URLS,
    _extract_bond_name,
    _serialize,
    _to_date,
    _to_decimal,
    _to_int,
    download_xlsx_files,
    parse_all,
    parse_calculator_xlsx,
    parse_indexed_xlsx,
    parse_prices_xlsx,
)


class TestSerialization:
    def test_serialize_none(self):
        assert _serialize(None) is None

    def test_serialize_datetime(self):
        assert _serialize(datetime(2026, 1, 15, 12, 30)) == "2026-01-15"

    def test_serialize_date(self):
        assert _serialize(date(2026, 2, 3)) == "2026-02-03"

    def test_serialize_numbers(self):
        assert _serialize(42) == 42
        assert _serialize(4.5) == 4.5

    def test_serialize_other(self):
        assert _serialize(Decimal("1.5")) == "1.5"
        assert _serialize("abc") == "abc"


class TestToDecimal:
    def test_none_and_empty(self):
        assert _to_decimal(None) is None
        assert _to_decimal("") is None
        assert _to_decimal("None") is None

    def test_decimal_passthrough(self):
        v = Decimal("12.34")
        assert _to_decimal(v) is v

    def test_numbers(self):
        assert _to_decimal(12) == Decimal("12")
        assert _to_decimal(12.5) == Decimal("12.5")

    def test_string_cleaning(self):
        assert _to_decimal("1 234,56%") == Decimal("1234.56")
        assert _to_decimal("10,5") == Decimal("10.5")

    def test_garbage(self):
        assert _to_decimal("abc") is None
        assert _to_decimal([1, 2]) is None


class TestToDate:
    def test_none_and_empty(self):
        assert _to_date(None) is None
        assert _to_date("") is None

    def test_date_and_datetime(self):
        d = date(2026, 1, 1)
        assert _to_date(d) is d
        assert _to_date(datetime(2026, 1, 1, 10, 0)) == d

    def test_iso_string(self):
        assert _to_date("2026-01-02") == date(2026, 1, 2)

    def test_dotted_string(self):
        assert _to_date("02.01.2026") == date(2026, 1, 2)

    def test_invalid(self):
        assert _to_date("not-a-date") is None
        assert _to_date(123) is None


class TestToInt:
    def test_none_and_empty(self):
        assert _to_int(None) is None
        assert _to_int("") is None
        assert _to_int("None") is None

    def test_numbers(self):
        assert _to_int(7) == 7
        assert _to_int(7.9) == 7

    def test_strings(self):
        assert _to_int("42") == 42
        assert _to_int(" 12.9 ") == 12

    def test_garbage_extracts_digits(self):
        assert _to_int("выпуск 17") == 17

    def test_no_digits(self):
        assert _to_int("abc") is None
        assert _to_int(object()) is None

    def test_overflow_string(self):
        assert _to_int("1e999") == 1


class TestExtractBondName:
    def test_with_digits(self):
        assert _extract_bond_name(" Айгенис 23 ") == ("Айгенис 23", 23)

    def test_without_digits(self):
        assert _extract_bond_name("Облигация") == ("Облигация", None)


def _make_calculator_workbook(tmp_path, sheets=None):
    import openpyxl

    wb = openpyxl.Workbook()
    default = wb.active
    default.title = "Мусор"
    for i in range(1, 6):
        wb.create_sheet(f"Лист{i}")
    sheets = sheets or {
        "Айгенис 23": {
            3: 1000.0,
            4: 10,
            5: 1000000.0,
            6: 12.5,
            7: date(2026, 1, 15),
            8: date(2031, 1, 15),
            9: 1825,
        },
        "Айгенис 24": {
            3: 1000.0,
            4: 5,
            5: 500000.0,
            6: 10.0,
            7: date(2026, 3, 1),
            8: date(2030, 3, 1),
            9: 1461,
        },
    }
    master = wb["Лист3"]
    master.cell(2, 2, "Характеристика")
    for col_idx, (name, params) in enumerate(sheets.items()):
        c = 4 + col_idx * 2
        master.cell(2, c, name)
        for row, value in params.items():
            master.cell(row, c, value)
    for name in sheets:
        ws = wb.create_sheet(f"{name} купоны")
        ws.cell(2, 1, 1)
        ws.cell(2, 2, date(2026, 1, 15))
        ws.cell(2, 3, date(2026, 7, 15))
        ws.cell(2, 4, 181)
        ws.cell(2, 5, 61.99)
        ws.cell(3, 1, 2)
        ws.cell(3, 2, date(2026, 7, 15))
        ws.cell(3, 3, date(2027, 1, 15))
        ws.cell(3, 4, 184)
        ws.cell(3, 5, 63.01)
    path = tmp_path / "calculator.xlsx"
    wb.save(str(path))
    return path


class TestParseCalculator:
    def test_parses_bonds_and_params(self, tmp_path):
        path = _make_calculator_workbook(tmp_path)
        bonds = parse_calculator_xlsx(str(path))
        assert set(bonds) == {23, 24}
        bond = bonds[23]
        assert bond.issue_number == 23
        assert bond.name == "Айгенис 23"
        assert bond.face_value == Decimal("1000")
        assert bond.quantity == 10
        assert bond.issue_volume == Decimal("1000000")
        assert bond.coupon_rate == Decimal("12.5")
        assert bond.start_date == date(2026, 1, 15)
        assert bond.maturity_date == date(2031, 1, 15)
        assert bond.term_days == 1825

    def test_coupon_schedule(self, tmp_path):
        path = _make_calculator_workbook(tmp_path)
        bond = parse_calculator_xlsx(str(path))[23]
        assert len(bond.coupon_periods) == 2
        assert bond.coupon_periods[0] == {
            "num": 1,
            "start": "2026-01-15",
            "end": "2026-07-15",
            "days": 181,
            "amount": 61.99,
        }
        assert bond.coupon_periods[1]["num"] == 2
        assert bond.coupon_periods[1]["amount"] == 63.01

    def test_coupon_schedule_skips_empty_rows(self, tmp_path):
        import openpyxl

        path = _make_calculator_workbook(tmp_path)
        wb = openpyxl.load_workbook(str(path))
        ws = wb["Айгенис 23 купоны"]
        ws.cell(4, 1, None)
        ws.cell(5, 1, 3)
        ws.cell(5, 2, date(2027, 1, 15))
        ws.cell(5, 3, date(2027, 7, 15))
        wb.save(str(path))
        bond = parse_calculator_xlsx(str(path))[23]
        assert [p["num"] for p in bond.coupon_periods] == [1, 2, 3]

    def test_aigen18_rf_maps_to_18(self, tmp_path):
        path = _make_calculator_workbook(tmp_path, sheets={"Aigen18-RF": {3: 100.0, 4: 1, 6: 5.0}})
        bonds = parse_calculator_xlsx(str(path))
        assert set(bonds) == {18}
        assert bonds[18].name == "Aigen18-RF"
        assert bonds[18].issue_number == 18

    def test_skips_names_without_issue_number(self, tmp_path):
        path = _make_calculator_workbook(tmp_path, sheets={"Облигация без номера": {3: 1.0}})
        bonds = parse_calculator_xlsx(str(path))
        assert bonds == {}

    def test_skips_empty_and_header_columns(self, tmp_path):
        path = _make_calculator_workbook(tmp_path, sheets={"Айгенис 7": {3: 100.0}})
        import openpyxl

        wb = openpyxl.load_workbook(str(path))
        master = wb["Лист3"]
        master.cell(2, 8, "Характеристика выпуска")
        master.cell(2, 10, " ")
        wb.save(str(path))
        bonds = parse_calculator_xlsx(str(path))
        assert set(bonds) == {7}

    def test_parse_all_single_file(self, tmp_path):
        path = _make_calculator_workbook(tmp_path)
        result = parse_all(str(path))
        assert set(result.byn_bonds) == {23, 24}
        assert result.indexed_bonds == {}
        assert result.daily_accruals == []

    def test_parse_all_directory(self, tmp_path):
        path = _make_calculator_workbook(tmp_path)
        dest = tmp_path / "dir"
        dest.mkdir()
        target = dest / os.path.basename(XLSX_URLS["calculator"])
        os.replace(str(path), str(target))
        result = parse_all(str(dest))
        assert set(result.byn_bonds) == {23, 24}

    def test_parse_all_full_directory(self, tmp_path):
        calc = _make_calculator_workbook(tmp_path)
        indexed = _make_indexed_workbook(tmp_path)
        prices = _make_prices_workbook(tmp_path, {23: [(date(2026, 6, 1), 1.0, 2.0)]})
        dest = tmp_path / "full"
        dest.mkdir()
        for path, key in (
            (calc, "calculator"),
            (indexed, "indexed"),
            (prices, "prices"),
        ):
            os.replace(str(path), str(dest / os.path.basename(XLSX_URLS[key])))
        result = parse_all(str(dest))
        assert set(result.byn_bonds) == {23, 24}
        assert set(result.indexed_bonds) == {"Оп17", "Оп18"}
        assert len(result.daily_accruals) == 1
        assert result.daily_accruals[0].internal_id == "23"

    def test_parse_all_ignores_other_files(self, tmp_path):
        calc = _make_calculator_workbook(tmp_path)
        dest = tmp_path / "mixed"
        dest.mkdir()
        (dest / "readme.txt").write_text("not an xlsx")
        (dest / "other.xlsx").write_bytes(b"not matched")
        os.replace(str(calc), str(dest / os.path.basename(XLSX_URLS["calculator"])))
        result = parse_all(str(dest))
        assert set(result.byn_bonds) == {23, 24}

    def test_parse_all_default_downloads(self, tmp_path, monkeypatch):
        calc = _make_calculator_workbook(tmp_path)
        import scraper.parsers.xlsx as xlsx_mod

        monkeypatch.setattr(xlsx_mod, "download_xlsx_files", lambda dest_dir=None: {"calculator": str(calc)})
        result = parse_all()
        assert set(result.byn_bonds) == {23, 24}
        assert result.daily_accruals == []
        assert result.indexed_bonds == {}


def _make_indexed_workbook(tmp_path):
    import openpyxl

    wb = openpyxl.Workbook()
    wb.active.title = "Лист0"
    for i in range(1, 7):
        wb.create_sheet(f"Лист{i}")
    master = wb["Лист5"]
    master.cell(2, 2, "Характеристика")
    master.cell(2, 4, "Айгенис Оп17_BYN→USD")
    master.cell(2, 6, "Айгенис Оп18_BYN→USD")
    master.cell(2, 8, "Без операционного номера")
    params = {
        3: 1000.0,
        4: 2.5,
        5: "usd",
        6: 20,
        7: 2000000.0,
        8: 6.5,
        10: date(2026, 5, 1),
        11: date(2031, 5, 1),
        12: 1826,
    }
    for col in (4, 6):
        for row, value in params.items():
            master.cell(row, col, value)
    for op in (17, 18):
        ws = wb.create_sheet(f"Оп{op}_BYN→USD")
        ws.cell(2, 1, 1)
        ws.cell(2, 2, date(2026, 5, 1))
        ws.cell(2, 3, date(2026, 11, 1))
        ws.cell(2, 4, 184)
        ws.cell(2, 5, 32.5)
    path = tmp_path / "indexed.xlsx"
    wb.save(str(path))
    return path


class TestParseIndexed:
    def test_parses_op_bonds(self, tmp_path):
        path = _make_indexed_workbook(tmp_path)
        bonds = parse_indexed_xlsx(str(path))
        assert set(bonds) == {"Оп17", "Оп18"}
        bond = bonds["Оп17"]
        assert bond.issue_number == 17
        assert bond.name == "Айгенис Оп17_BYN→USD"
        assert bond.face_value == Decimal("1000")
        assert bond.exchange_rate_on_start == Decimal("2.5")
        assert bond.indexation_currency == "USD"
        assert bond.quantity == 20
        assert bond.issue_volume == Decimal("2000000")
        assert bond.coupon_rate == Decimal("6.5")
        assert bond.start_date == date(2026, 5, 1)
        assert bond.maturity_date == date(2031, 5, 1)
        assert bond.term_days == 1826

    def test_uppercases_currency(self, tmp_path):
        path = _make_indexed_workbook(tmp_path)
        assert parse_indexed_xlsx(str(path))["Оп18"].indexation_currency == "USD"

    def test_skips_names_without_op_number(self, tmp_path):
        path = _make_indexed_workbook(tmp_path)
        bonds = parse_indexed_xlsx(str(path))
        assert "Оп0" not in bonds

    def test_coupon_schedule(self, tmp_path):
        path = _make_indexed_workbook(tmp_path)
        periods = parse_indexed_xlsx(str(path))["Оп17"].coupon_periods
        assert len(periods) == 1
        assert periods[0] == {
            "num": 1,
            "start": "2026-05-01",
            "end": "2026-11-01",
            "days": 184,
            "amount": 32.5,
        }

    def test_coupon_schedule_skips_empty_rows(self, tmp_path):
        import openpyxl

        path = _make_indexed_workbook(tmp_path)
        wb = openpyxl.load_workbook(str(path))
        ws = wb["Оп17_BYN→USD"]
        ws.cell(3, 1, "")
        wb.save(str(path))
        periods = parse_indexed_xlsx(str(path))["Оп17"].coupon_periods
        assert len(periods) == 1


def _make_prices_workbook(tmp_path, byn_bonds):
    import openpyxl

    wb = openpyxl.Workbook()
    wb.active.title = "Справочник"
    for issue, rows in byn_bonds.items():
        ws = wb.create_sheet(f"Текущая стоимость {issue}")
        for r, (d, accrued, total) in enumerate(rows, start=2):
            ws.cell(r, 1, d)
            ws.cell(r, 2, accrued)
            ws.cell(r, 3, total)
    ws = wb.create_sheet("Текущая стоимость без номера")
    ws.cell(2, 1, date(2026, 6, 1))
    ws.cell(2, 2, 1.0)
    path = tmp_path / "prices.xlsx"
    wb.save(str(path))
    return path


class TestParsePrices:
    def test_parses_accrual_rows(self, tmp_path):
        path = _make_prices_workbook(
            tmp_path, {23: [(date(2026, 6, 1), 100.5, 1100.5), (date(2026, 6, 2), 101.0, 1101.0)]}
        )
        accruals = parse_prices_xlsx(str(path), {23: object()})
        assert len(accruals) == 2
        assert accruals[0].internal_id == "23"
        assert accruals[0].date == date(2026, 6, 1)
        assert accruals[0].accrued == Decimal("100.5")
        assert accruals[0].total_value == Decimal("1100.5")

    def test_skips_sheet_without_digits(self, tmp_path):
        path = _make_prices_workbook(tmp_path, {23: [(date(2026, 6, 1), 1.0, 2.0)]})
        accruals = parse_prices_xlsx(str(path), {})
        assert accruals == []

    def test_skips_issue_not_in_enrichment(self, tmp_path):
        path = _make_prices_workbook(tmp_path, {99: [(date(2026, 6, 1), 1.0, 2.0)]})
        accruals = parse_prices_xlsx(str(path), {23: object()})
        assert accruals == []

    def test_skips_rows_without_date(self, tmp_path):
        path = _make_prices_workbook(
            tmp_path, {23: [(None, 1.0, 2.0), (date(2026, 6, 1), 3.0, 4.0)]}
        )
        accruals = parse_prices_xlsx(str(path), {23: object()})
        assert len(accruals) == 1
        assert accruals[0].date == date(2026, 6, 1)


class TestDownload:
    def test_downloads_and_caches(self, tmp_path, monkeypatch):
        calls = []
        original_urlopen = __import__("urllib.request", fromlist=["urlopen"]).urlopen

        class FakeResp:
            def __init__(self, data):
                self._data = data

            def __enter__(self):
                return self

            def __exit__(self, *args):
                return False

            def read(self):
                return self._data

        def fake_urlopen(url, timeout=None):
            calls.append(url)
            return FakeResp(b"fake-xlsx-content")

        monkeypatch.setattr(original_urlopen.__module__ + ".urlopen", fake_urlopen)
        result = download_xlsx_files(str(tmp_path))
        assert set(result) == set(XLSX_URLS)
        assert len(calls) == 3
        for key, path in result.items():
            assert os.path.exists(path)
            with open(path, "rb") as f:
                assert f.read() == b"fake-xlsx-content"

        result2 = download_xlsx_files(str(tmp_path))
        assert len(calls) == 3

    def test_download_into_temp_default(self, tmp_path, monkeypatch):
        monkeypatch.setenv("TEMP", str(tmp_path))

        class FakeResp:
            def __enter__(self):
                return self

            def __exit__(self, *args):
                return False

            def read(self):
                return b"x"

        monkeypatch.setattr("urllib.request.urlopen", lambda url, timeout: FakeResp())
        result = download_xlsx_files()
        assert os.path.exists(result["calculator"])
