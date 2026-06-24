from __future__ import annotations

import os
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from scraper.models import BondDailyAccrual

XLSX_URLS = {
    "prices": "https://aigenis.by/wp-content/uploads/2026/06/tekushhaya_stoimost_obligaczij_ajgenis_byn_sajt.xlsx",
    "calculator": "https://aigenis.by/wp-content/uploads/2026/04/kalkulyator_investora_obligaczii_aigenis_24.04.2026.xlsx",
    "indexed": "https://aigenis.by/wp-content/uploads/2026/04/kalkulyator_investora_indeksiruemye_obligaczii_aigenis_24_04_2026.xlsx",
}


@dataclass
class BondXlsxEnrichment:
    issue_number: int | None = None
    name: str | None = None
    face_value: Decimal | None = None
    quantity: int | None = None
    issue_volume: Decimal | None = None
    coupon_rate: Decimal | None = None
    start_date: date | None = None
    maturity_date: date | None = None
    term_days: int | None = None
    indexation_currency: str | None = None
    exchange_rate_on_start: Decimal | None = None
    coupon_periods: list[dict] = field(default_factory=list)
    daily_accruals: list[BondDailyAccrual] = field(default_factory=list)


@dataclass
class XlsxParseResult:
    byn_bonds: dict[int, BondXlsxEnrichment] = field(default_factory=dict)
    indexed_bonds: dict[str, BondXlsxEnrichment] = field(default_factory=dict)
    daily_accruals: list[BondDailyAccrual] = field(default_factory=list)


def _serialize(v: Any) -> str | float | int | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, (int, float)):
        return v
    return str(v)


def _to_decimal(v: Any) -> Decimal | None:
    if v is None or v == "" or v == "None":
        return None
    if isinstance(v, Decimal):
        return v
    if isinstance(v, (int, float)):
        return Decimal(str(v))
    if isinstance(v, str):
        try:
            return Decimal(v.strip().replace(",", "."))
        except Exception:
            return None
    return None


def _to_date(v: Any) -> date | None:
    if v is None or v == "":
        return None
    if isinstance(v, date):
        return v
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, str):
        for fmt in ("%Y-%m-%d", "%d.%m.%Y"):
            try:
                return datetime.strptime(v[:10], fmt).date()
            except ValueError:
                continue
    return None


def _to_int(v: Any) -> int | None:
    if v is None or v == "" or v == "None":
        return None
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        return int(v)
    if isinstance(v, str):
        try:
            return int(float(v.strip()))
        except (ValueError, OverflowError):
            import re
            m = re.search(r"\d+", v)
            return int(m.group()) if m else None
    return None


def _load_workbook(path: str):
    import openpyxl
    return openpyxl.load_workbook(path, data_only=True)


def _extract_bond_name(name: str) -> tuple[str | None, int | None]:
    name = name.strip()
    m = __import__("re").search(r"(\d+)", name)
    issue_num = int(m.group(1)) if m else None
    return name, issue_num


def parse_calculator_xlsx(filepath: str) -> dict[int, BondXlsxEnrichment]:
    """Parse the investor calculator XLSX for BYN bonds."""
    wb = _load_workbook(filepath)
    sheets = wb.sheetnames

    master_ws = wb[sheets[3]]

    bonds: dict[int, BondXlsxEnrichment] = {}

    # Layout: col 2 = label, cols 4,6,8,... = bond data columns
    # Row 2: bond names in data columns
    bond_cols = []
    for c in range(4, master_ws.max_column + 1, 2):
        name = str(master_ws.cell(2, c).value or "").strip()
        if not name or name in ("", " ", "None", "Характеристика", "Характеристикавыпуска", "Характеристика выпуска"):
            continue
        bond_cols.append((c, name))

    param_row_map = {
        3: "face_value",
        4: "quantity",
        5: "issue_volume",
        6: "coupon_rate",
        7: "start_date",
        8: "maturity_date",
        9: "term_days",
    }

    for col_idx, bond_name in bond_cols:
        _, issue_num = _extract_bond_name(bond_name)
        if issue_num is None and bond_name != "Aigen18-RF":
            continue

        key = 18 if bond_name == "Aigen18-RF" else issue_num
        bond = BondXlsxEnrichment(issue_number=key, name=bond_name)

        for r, field_name in param_row_map.items():
            raw_val = master_ws.cell(r, col_idx).value
            if field_name in ("face_value", "issue_volume", "coupon_rate"):
                setattr(bond, field_name, _to_decimal(raw_val))
            elif field_name in ("quantity", "term_days"):
                setattr(bond, field_name, _to_int(raw_val))
            elif field_name in ("start_date", "maturity_date"):
                setattr(bond, field_name, _to_date(raw_val))

        # Parse coupon schedule from individual bond sheet (match by issue number)
        for sn in sheets:
            if key is not None and str(key) in sn:
                ws = wb[sn]
                periods = []
                for r in range(2, ws.max_row + 1):
                    num = ws.cell(r, 1).value
                    if num is None or num == "":
                        continue
                    period_start = _to_date(ws.cell(r, 2).value)
                    period_end = _to_date(ws.cell(r, 3).value)
                    days = _to_int(ws.cell(r, 4).value)
                    amount = _to_decimal(ws.cell(r, 5).value)
                    if period_start and period_end:
                        periods.append({
                            "num": _to_int(num),
                            "start": str(period_start),
                            "end": str(period_end),
                            "days": days,
                            "amount": float(amount) if amount else None,
                        })
                if periods:
                    bond.coupon_periods = periods
                break

        bonds[key] = bond

    wb.close()
    return bonds


def parse_indexed_xlsx(filepath: str) -> dict[str, BondXlsxEnrichment]:
    """Parse the indexed bonds calculator XLSX."""
    wb = _load_workbook(filepath)
    sheets = wb.sheetnames

    # Find master sheet (Все выпуски, index 5)
    master_ws = wb[sheets[5]]

    bonds: dict[str, BondXlsxEnrichment] = {}

    param_rows = {}
    for r in range(3, 12):
        param_name = _serialize(master_ws.cell(r, 1).value)
        if not param_name:
            continue
        values = []
        for c in range(2, master_ws.max_column + 1):
            values.append(master_ws.cell(c if c % 2 == 0 else c, r).value)

    # Better approach: read the master table more carefully
    bond_names = []
    for c in range(2, master_ws.max_column + 1, 2):
        name = _serialize(master_ws.cell(2, c).value)
        if name and name not in ("", " ", "Характеристика", "Характеристикавыпуска", "Характеристика выпуска"):
            bond_names.append((c, name))

    param_map = {
        3: ("face_value", "decimal"),
        4: ("exchange_rate_on_start", "decimal"),
        5: ("indexation_currency", "str"),
        6: ("quantity", "int"),
        7: ("issue_volume", "decimal"),
        8: ("coupon_rate", "decimal"),
        10: ("start_date", "date"),
        11: ("maturity_date", "date"),
        12: ("term_days", "int"),
    }

    for col_idx, bond_name in bond_names:
        # Extract op number (e.g., "Оп17" from "Айгенис Оп17_BYN→USD")
        import re
        m = re.search(r"Оп(\d+)", bond_name)
        op_num = m.group(1) if m else None
        if not op_num:
            continue

        key = f"Оп{op_num}"
        bond = BondXlsxEnrichment(issue_number=int(op_num), name=bond_name, indexation_currency="USD")

        for r in range(3, 13):
            raw_val = master_ws.cell(r, col_idx).value
            field_name, ftype = param_map.get(r, (None, None))
            if not field_name:
                continue
            if ftype == "decimal":
                setattr(bond, field_name, _to_decimal(raw_val))
            elif ftype == "int":
                setattr(bond, field_name, _to_int(raw_val))
            elif ftype == "date":
                setattr(bond, field_name, _to_date(raw_val))
            elif ftype == "str":
                val = _serialize(raw_val)
                if val and val not in ("", " ", "None"):
                    if field_name == "indexation_currency":
                        val = val.upper()
                    setattr(bond, field_name, val)

        # Parse coupon schedule from individual bond sheet
        for sn in sheets:
            if sn.startswith("Оп"):
                # sheet names like "Оп17_BYN→USD"
                sheet_op = sn.split("_")[0] if "_" in sn else sn
                if sheet_op == key:
                    ws = wb[sn]
                    periods = []
                    for r in range(2, ws.max_row + 1):
                        num = ws.cell(r, 1).value
                        if num is None or num == "":
                            continue
                        period_start = _to_date(ws.cell(r, 2).value)
                        period_end = _to_date(ws.cell(r, 3).value)
                        days = _to_int(ws.cell(r, 4).value)
                        amount = _to_decimal(ws.cell(r, 5).value)
                        if period_start and period_end:
                            periods.append({
                                "num": _to_int(num),
                                "start": str(period_start),
                                "end": str(period_end),
                                "days": days,
                                "amount": float(amount) if amount else None,
                            })
                    if periods:
                        bond.coupon_periods = periods
                    break

        bonds[key] = bond

    wb.close()
    return bonds


def parse_prices_xlsx(filepath: str, byn_bonds: dict[int, BondXlsxEnrichment]) -> list[BondDailyAccrual]:
    """Parse the current prices XLSX for daily accrual data."""
    wb = _load_workbook(filepath)
    sheets = wb.sheetnames

    all_accruals: list[BondDailyAccrual] = []

    for sn in sheets:
        if "Текущ" not in sn:
            continue

        # Extract issue number from sheet name
        import re
        m = re.search(r"(\d+)", sn)
        if not m:
            continue
        issue_num = int(m.group(1))

        # Map to internal_id from enrichment data
        bond_key = issue_num
        enrichment = byn_bonds.get(bond_key)
        if not enrichment:
            continue

        ws = wb[sn]
        for r in range(2, ws.max_row + 1):
            d = _to_date(ws.cell(r, 1).value)
            accrued = _to_decimal(ws.cell(r, 2).value)
            total = _to_decimal(ws.cell(r, 3).value)
            if d is None:
                continue

            # We don't have the internal_id here, just issue_number
            all_accruals.append(BondDailyAccrual(
                internal_id=str(issue_num),
                date=d,
                accrued=accrued,
                total_value=total,
            ))

    wb.close()
    return all_accruals


def download_xlsx_files(dest_dir: str | None = None) -> dict[str, str]:
    """Download all XLSX files to a local directory. Returns dict of name -> local path."""
    if dest_dir is None:
        dest_dir = os.environ.get("TEMP", "/tmp")
    os.makedirs(dest_dir, exist_ok=True)

    import urllib.request

    result = {}
    for key, url in XLSX_URLS.items():
        local_path = os.path.join(dest_dir, os.path.basename(url))
        if not os.path.exists(local_path):
            urllib.request.urlretrieve(url, local_path)
        result[key] = local_path
    return result


def parse_all(filepath_or_dir: str | None = None) -> XlsxParseResult:
    """Parse all XLSX files and return structured data."""
    if filepath_or_dir is None:
        paths = download_xlsx_files()
    elif os.path.isdir(filepath_or_dir):
        paths = {}
        for fname in os.listdir(filepath_or_dir):
            if fname.endswith(".xlsx"):
                for key, url in XLSX_URLS.items():
                    if os.path.basename(url) == fname:
                        paths[key] = os.path.join(filepath_or_dir, fname)
    else:
        paths = {"calculator": filepath_or_dir}

    result = XlsxParseResult()

    if "calculator" in paths:
        result.byn_bonds = parse_calculator_xlsx(paths["calculator"])

    if "prices" in paths:
        result.daily_accruals = parse_prices_xlsx(paths["prices"], result.byn_bonds)

    if "indexed" in paths:
        result.indexed_bonds = parse_indexed_xlsx(paths["indexed"])

    return result
